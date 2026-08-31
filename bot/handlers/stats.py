"""
هندلرهای بخش آمار و گزارش برای ادمین:
- نمایش آمار کلی (کاربران، رویدادها، وضعیت ثبت‌نام‌ها، نشریه، FAQ)
- خروجی اکسل لیست ثبت‌نامی‌های هر رویداد (شامل جواب فیلدهای پویا)
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from telegram import InlineKeyboardButton, InlineKeyboardMarkup, Update
from telegram.ext import ContextTypes

from bot.handlers.events import STATUS_LABELS
from bot.handlers.start import is_admin_telegram_id
from database.db import get_session
from database.models import (
    Admin,
    Event,
    EventField,
    FAQ,
    JournalIssue,
    Registration,
    RegistrationFieldValue,
    RegistrationStatus,
    UserAccount,
)


async def admin_stats_menu_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()
    if not is_admin_telegram_id(query.from_user.id):
        await query.edit_message_text("⛔️ شما به این بخش دسترسی ندارید.")
        return

    session = get_session()
    try:
        total_users = session.query(UserAccount).count()
        total_admins = session.query(Admin).count()
        total_events = session.query(Event).count()
        active_events = session.query(Event).filter_by(is_active=True).count()
        total_journal = session.query(JournalIssue).count()
        total_faq = session.query(FAQ).count()

        status_counts = {}
        for status in RegistrationStatus:
            status_counts[status] = session.query(Registration).filter_by(status=status).count()
    finally:
        session.close()

    lines = [
        "📊 آمار کلی ربات:",
        "",
        f"👤 کاربران: {total_users}",
        f"👥 ادمین‌ها: {total_admins}",
        f"📅 رویدادها: {total_events} (فعال: {active_events})",
        f"📚 شماره‌های نشریه: {total_journal}",
        f"❓ سوالات متداول: {total_faq}",
        "",
        "📋 ثبت‌نام‌ها به تفکیک وضعیت:",
    ]
    for status, count in status_counts.items():
        lines.append(f"  • {STATUS_LABELS.get(status, status)}: {count}")

    buttons = [
        [InlineKeyboardButton("📥 خروجی اکسل ثبت‌نامی‌های یک رویداد", callback_data="admin_stats_pick_event")],
        [InlineKeyboardButton("⬅️ بازگشت", callback_data="admin_menu")],
    ]
    await query.edit_message_text("\n".join(lines), reply_markup=InlineKeyboardMarkup(buttons))


async def admin_stats_pick_event_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()
    if not is_admin_telegram_id(query.from_user.id):
        await query.edit_message_text("⛔️ شما به این بخش دسترسی ندارید.")
        return

    session = get_session()
    try:
        events = session.query(Event).order_by(Event.created_at.desc()).all()
        rows = [(e.id, e.title) for e in events]
    finally:
        session.close()

    if not rows:
        keyboard = InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ بازگشت", callback_data="admin_stats")]])
        await query.edit_message_text("هنوز هیچ رویدادی ثبت نشده است.", reply_markup=keyboard)
        return

    buttons = [[InlineKeyboardButton(title, callback_data=f"admin_stats_export_{eid}")] for eid, title in rows]
    buttons.append([InlineKeyboardButton("⬅️ بازگشت", callback_data="admin_stats")])
    await query.edit_message_text(
        "کدام رویداد را می‌خواهید خروجی بگیرید؟", reply_markup=InlineKeyboardMarkup(buttons)
    )


def _build_registrations_excel(event_id: str):
    session = get_session()
    try:
        event = session.query(Event).filter_by(id=event_id).first()
        if event is None:
            return None, None

        fields = (
            session.query(EventField)
            .filter_by(event_id=event_id)
            .order_by(EventField.display_order)
            .all()
        )
        registrations = (
            session.query(Registration)
            .filter_by(event_id=event_id)
            .order_by(Registration.submitted_at)
            .all()
        )

        has_price = bool(event.price)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "ثبت‌نام‌ها"

        headers = ["کد پیگیری", "کاربر تلگرام", "وضعیت"] + [f.field_label for f in fields]
        if has_price:
            headers.append("مبلغ پرداختی (تومان)")
        headers.append("تاریخ ثبت‌نام")
        sheet.append(headers)

        total_approved = 0
        for reg in registrations:
            user = session.query(UserAccount).filter_by(id=reg.user_id).first()
            values = session.query(RegistrationFieldValue).filter_by(registration_id=reg.id).all()
            value_map = {v.event_field_id: v.value for v in values}

            user_label = "-"
            if user:
                user_label = f"@{user.username}" if user.username else (user.full_name or str(user.telegram_id))

            row = [reg.tracking_code, user_label, STATUS_LABELS.get(reg.status, reg.status)]
            row.extend(value_map.get(f.id, "") for f in fields)
            if has_price:
                # مبلغ فقط برای کسانی که واقعاً فیش فرستاده‌اند (یعنی رایگان نبوده) نمایش داده می‌شود
                row.append(int(event.price) if reg.receipt_file_url else "")
                if reg.status == RegistrationStatus.APPROVED:
                    total_approved += int(event.price)
            row.append(reg.submitted_at.strftime("%Y-%m-%d %H:%M") if reg.submitted_at else "")
            sheet.append(row)

        if has_price:
            amount_col_index = 3 + len(fields)  # ایندکس ستون «مبلغ پرداختی» در headers
            summary = ["" for _ in headers]
            summary[0] = "جمع مبلغ ثبت‌نامی‌های تاییدشده:"
            summary[amount_col_index] = total_approved
            sheet.append([])
            sheet.append(summary)

        for i in range(1, len(headers) + 1):
            sheet.column_dimensions[get_column_letter(i)].width = 22

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer, event.title
    finally:
        session.close()


async def admin_stats_export_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer("در حال آماده‌سازی فایل...")
    if not is_admin_telegram_id(query.from_user.id):
        await query.edit_message_text("⛔️ شما به این بخش دسترسی ندارید.")
        return

    event_id = context.match.group("event_id")
    buffer, title = _build_registrations_excel(event_id)
    if buffer is None:
        await context.bot.send_message(chat_id=query.from_user.id, text="این رویداد یافت نشد.")
        return

    await context.bot.send_document(
        chat_id=query.from_user.id,
        document=buffer,
        filename=f"{title}.xlsx",
        caption=f"📥 لیست ثبت‌نامی‌های «{title}»",
    )
